# Скрипт сопоставляет данные из двух файлов эксель и заносит их в третий файл.


import glob
import os
import datetime
import pandas as pd 
import openpyxl

from tkinter import filedialog
from tkinter import messagebox

#list_of_files = glob.glob(r'C:/xl/des/rn/load/*.xlsx') # * find last file
#latest_file = max(list_of_files, key=os.path.getctime)

filename = filedialog.askopenfilename(title = "Выберите файл обновления")
wb = openpyxl.load_workbook(filename)									 
ws = wb.active															 


now = datetime.datetime.now()	
td = now.strftime("%d.%m.%Y")		


filename2 = filedialog.askopenfilename(title = "Выберите файл выгрузки")
ld1 = openpyxl.load_workbook(filename2)									
ld = ld1.active															


a1, a2, a3, a4, a5, a6, a7, a8, a9, dt = [],[],[],[],[],[],[],[],[],[] 
prod = ws['L']
proc = ws['M']
htrf = ws['N']
gosp = ws['O'] 
ostk = ws['P']
empt = ws['NN']
dgv2 = ws['K']
dgv1 = ld['E']
iddg = ld['G']

for b1 in prod:		   
	a1.append(b1.value)
for b2 in proc:		   
	a2.append(b2.value)
for b3 in htrf: 	   
	a3.append(b3.value)
for b4 in gosp:		   
	a4.append(b4.value)
for b5 in ostk:		   
	a5.append(b5.value)
for b6 in dgv2:		   
	a6.append(b6.value)


for b7 in dgv1:			
	a7.append(b7.value)   
for b8 in iddg:				
	a8.append(str(b8.value))#	
for b9 in empt:				
	a9.append(b9.value)		
	dt.append(td)			

a6[0] = a7[0]				



pd1 = pd.DataFrame({a6[0]:a6[1:], a5[0]:a5[1:], 		
				'emp1':a9[1:], a1[0]:a1[1:],		
				'emp2':a9[1:], a2[0]:a2[1:], a3[0]:a3[1:], 
				'emp3':a9[1:], 								
				'emp4':a9[1:], a4[0]:a4[1:],  			
				'Дата расчета':dt[1:],				
				'emp5':a9[1:], 'emp6':a9[1:]})				

pd2 = pd.DataFrame({a7[0]:a7[1:], a8[0]:a8[1:]})			


vpr = pd1.merge(pd2, on=a6[0], how='inner')#Ф-ция сопоставления двух файлов
v = list(vpr.columns)
v = v[-1:] + v[:-1]	
vpr = vpr[v]		

bb = vpr.drop('Номер договора', 1)

сс = bb.style.set_properties(**{'background-color': 'yellow'})
сс.to_excel(r"C:\Users\Admin\Desktop\tst\rn\load\res\ " +str(td) + ".xlsx", index = False, header = False)

messagebox.showinfo("COMPLETE", "Итоговый файл лежит на рабочем столе в папке tst/РН/res")
